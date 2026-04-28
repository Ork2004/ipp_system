"""
Auto-detect Form 63 template structure.

Given an uploaded XLSX, scan the header zone and produce:
- column_mapping: which column letter holds each known data category
- data_start_row: first row where teacher data should be written
- detection_meta: human-readable trace of what matched, for UI preview

This lets admins upload a freshly-issued Form 63 template each year
without anyone editing code. The known categories are:

    row_number, teacher_name, position, semester,
    teaching_auditory, teaching_extraauditory,
    methodical, research, organizational_methodical,
    educational, qualification, social,
    total,
    hourly_auditory, hourly_extraauditory
"""

import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


HEADER_SCAN_ROWS = 25


def _norm(text) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text))
    s = s.lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _contains_any(text: str, *keywords: str) -> bool:
    return any(kw in text for kw in keywords)


def _scan_cells(ws):
    cells = []
    last_row = min(HEADER_SCAN_ROWS, ws.max_row or HEADER_SCAN_ROWS)
    for row in ws.iter_rows(min_row=1, max_row=last_row):
        for cell in row:
            if cell.value is None:
                continue
            text = _norm(cell.value)
            if not text:
                continue
            cells.append({
                "row": cell.row,
                "col": cell.column,
                "letter": get_column_letter(cell.column),
                "text": text,
                "raw": str(cell.value),
            })
    return cells


def _find_first(cells, predicate):
    for c in cells:
        if predicate(c):
            return c
    return None


def _column_span_in_row(ws, parent_cell):
    """
    Span of a parent header that visually covers several columns.

    First check merged ranges. If the parent cell is in a merge that
    extends to the right, use that. Otherwise walk to the right in the
    same row and stop at the next non-empty cell — that defines the span.
    """
    row = parent_cell["row"]
    col = parent_cell["col"]

    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return (mr.min_col, mr.max_col)

    max_col = ws.max_column or col
    end_col = col
    for next_col in range(col + 1, max_col + 1):
        v = ws.cell(row=row, column=next_col).value
        if v is not None and _norm(v):
            break
        end_col = next_col
    return (col, end_col)


def _find_subheader_in_span(cells, span, header_row, predicate):
    """
    Look for a sub-header (e.g. 'Аудиторная') in rows immediately
    below header_row, but only within the column span of a parent header.
    """
    min_col, max_col = span
    for c in cells:
        if c["row"] <= header_row:
            continue
        if c["row"] > header_row + 3:
            continue
        if not (min_col <= c["col"] <= max_col):
            continue
        if predicate(c):
            return c
    return None


def _find_data_start_row(ws, header_row: int, row_number_col: int | None) -> int:
    """
    First row after the header zone that looks like a data row.

    Heuristic: walk rows after header_row + 1 and pick the first one whose
    row_number column holds the integer 1, OR is the first row past the
    header where most of the row is empty (template not yet filled).
    """
    start_search = header_row + 1
    end_search = min(ws.max_row or start_search, start_search + 20)

    if row_number_col:
        for r in range(start_search, end_search + 1):
            v = ws.cell(row=r, column=row_number_col).value
            if isinstance(v, (int, float)) and int(v) == 1:
                return r

    return header_row + 4


def parse_form63_template(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active

    cells = _scan_cells(ws)
    if not cells:
        raise ValueError("Шаблон выглядит пустым — заголовки не найдены")

    detection_meta = {"sheet": ws.title, "matches": [], "warnings": []}
    mapping: dict[str, str] = {}

    def record(category: str, cell: dict | None, note: str = ""):
        if cell is None:
            detection_meta["warnings"].append(f"Не найдена колонка для: {category}")
            return
        mapping[category] = cell["letter"]
        detection_meta["matches"].append({
            "category": category,
            "cell": f"{cell['letter']}{cell['row']}",
            "text": cell["raw"],
            "note": note,
        })

    # Anchor: ФИО ППС
    fio = _find_first(cells, lambda c: "фио" in c["text"])
    if fio is None:
        raise ValueError("Не найден заголовок 'ФИО' — шаблон не похож на Форму 63")
    record("teacher_name", fio)
    header_row = fio["row"]
    detection_meta["header_row"] = header_row

    # Должность
    position = _find_first(
        cells,
        lambda c: c["row"] == header_row and "должност" in c["text"],
    ) or _find_first(cells, lambda c: "должност" in c["text"])
    record("position", position)

    # Семестр
    semester = _find_first(cells, lambda c: "семестр" in c["text"])
    record("semester", semester)

    # Single-category data columns.
    # Length checks rule out long approval/title blocks that may incidentally
    # contain "учебно-методического совета", "научно-методический совет", etc.
    record(
        "methodical",
        _find_first(
            cells,
            lambda c: c["row"] >= header_row
            and ("учебно-методическ" in c["text"] or "учебно методическ" in c["text"])
            and "работа" in c["text"]
            and len(c["text"]) < 80,
        ),
    )
    record(
        "research",
        _find_first(
            cells,
            lambda c: c["row"] >= header_row
            and ("научная" in c["text"] or "научно-исследов" in c["text"])
            and "работа" in c["text"]
            and len(c["text"]) < 80,
        ),
    )
    record(
        "organizational_methodical",
        _find_first(
            cells,
            lambda c: c["row"] >= header_row
            and "организационно" in c["text"]
            and len(c["text"]) < 80,
        ),
    )
    record(
        "educational",
        _find_first(
            cells,
            lambda c: c["row"] >= header_row
            and ("воспитательная" in c["text"] or "профориентационная" in c["text"]),
        ),
    )
    # 'повышение квалификации' is a sub-phrase inside the educational header,
    # so look for it as a standalone short cell or by exact-ish match.
    qualification = _find_first(
        cells,
        lambda c: c["text"] == "повышение квалификации"
        or (
            "повышение квалификаци" in c["text"]
            and "воспитательная" not in c["text"]
            and len(c["text"]) < 60
        ),
    )
    record("qualification", qualification)

    social = _find_first(
        cells,
        lambda c: c["text"] == "общественная работа"
        or (
            "общественная работа" in c["text"]
            and "воспитательная" not in c["text"]
            and len(c["text"]) < 60
        ),
    )
    record("social", social)

    record(
        "total",
        _find_first(cells, lambda c: "итого" in c["text"]),
    )

    # Parent headers: regular teaching vs hourly
    teaching_parents = [
        c for c in cells
        if "учебная работа" in c["text"] and "почасов" not in c["text"]
    ]
    hourly_parents = [
        c for c in cells if "почасов" in c["text"]
    ]

    teaching_parent = teaching_parents[0] if teaching_parents else None
    hourly_parent = hourly_parents[0] if hourly_parents else None

    def find_in_parent(parent, keyword):
        if parent is None:
            return None
        span = _column_span_in_row(ws, parent)
        return _find_subheader_in_span(
            cells, span, parent["row"],
            lambda c: keyword in c["text"],
        )

    record("teaching_auditory", find_in_parent(teaching_parent, "аудиторная"),
           note="внутри 'Учебная работа'" if teaching_parent else "")
    record(
        "teaching_extraauditory",
        find_in_parent(teaching_parent, "внеаудиторная"),
        note="внутри 'Учебная работа'" if teaching_parent else "",
    )
    record("hourly_auditory", find_in_parent(hourly_parent, "аудиторная"),
           note="внутри 'Учебная работа (почасовая)'" if hourly_parent else "")
    record(
        "hourly_extraauditory",
        find_in_parent(hourly_parent, "внеаудиторная"),
        note="внутри 'Учебная работа (почасовая)'" if hourly_parent else "",
    )

    # Row number column (№). Often unlabeled but always to the left of ФИО.
    # If the cell directly to the left of ФИО header is empty/numeric, use it.
    row_number_letter = None
    if fio["col"] > 1:
        candidate_col = fio["col"] - 1
        # try a few candidates to the left
        for delta in range(1, 4):
            col_idx = fio["col"] - delta
            if col_idx < 1:
                break
            # column letter from data area
            row_number_letter = get_column_letter(col_idx)
            mapping["row_number"] = row_number_letter
            break

    # Data start row
    row_number_col = None
    if "row_number" in mapping:
        from openpyxl.utils import column_index_from_string
        row_number_col = column_index_from_string(mapping["row_number"])
    data_start_row = _find_data_start_row(ws, header_row, row_number_col)
    detection_meta["data_start_row"] = data_start_row

    return {
        "column_mapping": mapping,
        "data_start_row": data_start_row,
        "detection_meta": detection_meta,
    }


# Categories required for a usable Form 63 generation.
# If any are missing, the parser still returns, but the API marks the
# template as needing manual mapping override.
REQUIRED_CATEGORIES = (
    "teacher_name",
    "position",
    "semester",
    "teaching_auditory",
)


def missing_required(mapping: dict[str, str]) -> list[str]:
    return [c for c in REQUIRED_CATEGORIES if c not in mapping]
