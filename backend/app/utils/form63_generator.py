from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


# Order matters here only for the simple xlsx export.
DATA_CATEGORY_TO_FIELD = {
    "teaching_auditory": "teaching_auditory_hours",
    "teaching_extraauditory": "teaching_extraauditory_hours",
    "methodical": "methodical_hours",
    "research": "research_hours",
    "organizational_methodical": "organizational_methodical_hours",
    "educational": "educational_hours",
    "qualification": "qualification_hours",
    "social": "social_hours",
    "hourly_auditory": "hourly_auditory_hours",
    "hourly_extraauditory": "hourly_extraauditory_hours",
}


def build_form63_rows_from_preview(items: list[dict]) -> list[dict]:
    rows = []

    for item in items:
        rows.append({
            "teacher_name": item["teacher_name"],
            "position": item["position"],
            "semester": item["semester"],
            "teaching_auditory_hours": item.get("teaching_auditory_hours", 0),
            "teaching_extraauditory_hours": item.get("teaching_extraauditory_hours", 0),
            "methodical_hours": item.get("methodical_hours", 0),
            "research_hours": item.get("research_hours", 0),
            "organizational_methodical_hours": item.get("organizational_methodical_hours", 0),
            "educational_hours": item.get("educational_hours", 0),
            "qualification_hours": item.get("qualification_hours", 0),
            "social_hours": item.get("social_hours", 0),
            "planned_total_hours": item.get("planned_total_hours", 0),
            "hourly_auditory_hours": item.get("hourly_auditory_hours", 0),
            "hourly_extraauditory_hours": item.get("hourly_extraauditory_hours", 0),
        })

    return rows


def export_form63_simple_xlsx(rows: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Form63"

    headers = [
        "ФИО ППС",
        "Должность",
        "Семестр",
        "Учебная работа (аудиторная)",
        "Учебная работа (внеаудиторная)",
        "Учебно-методическая работа",
        "Научная работа",
        "Организационно-методическая работа",
        "Воспитательная работа",
        "Повышение квалификации",
        "Общественная работа",
        "Итого плановых часов",
        "Почасовая аудиторная работа",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([
            row["teacher_name"],
            row["position"],
            row["semester"],
            row["teaching_auditory_hours"],
            row["teaching_extraauditory_hours"],
            row["methodical_hours"],
            row["research_hours"],
            row["organizational_methodical_hours"],
            row["educational_hours"],
            row["qualification_hours"],
            row["social_hours"],
            row["planned_total_hours"],
            row["hourly_auditory_hours"],
        ])

    widths = {
        "A": 35, "B": 25, "C": 10, "D": 22, "E": 24, "F": 24,
        "G": 18, "H": 28, "I": 20, "J": 22, "K": 18, "L": 18, "M": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------- template-driven export ----------

def _is_merged_range(ws, range_str: str) -> bool:
    return range_str in {str(r) for r in ws.merged_cells.ranges}


def _safe_merge(ws, range_str: str):
    if _is_merged_range(ws, range_str):
        return
    # Some templates may already merge a wider range that contains this one.
    # In that case openpyxl raises on overlap; we just skip silently.
    try:
        ws.merge_cells(range_str)
    except Exception:
        pass


def _write_value(ws, letter: str, row: int, value):
    ws[f"{letter}{row}"] = value


def _put(ws, mapping: dict[str, str], category: str, row: int, value):
    """Write `value` into the column mapped to `category`, if mapping has it."""
    letter = mapping.get(category)
    if not letter:
        return
    _write_value(ws, letter, row, value)


def _category_span_in_header(ws, mapping: dict[str, str], category: str, header_row: int):
    """
    If a category column belongs to a horizontally-merged header range
    (e.g. ФИО merged across D:F), return the (min_letter, max_letter) of that
    merged span. Used to also merge data cells across the same horizontal span.
    """
    letter = mapping.get(category)
    if not letter:
        return None
    col = column_index_from_string(letter)
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= header_row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            if mr.max_col > mr.min_col:
                return (get_column_letter(mr.min_col), get_column_letter(mr.max_col))
    return None


def export_form63_from_template(
    rows: list[dict],
    template_path: str,
    output_path: str,
    column_mapping: dict[str, str],
    data_start_row: int,
):
    """
    Fill a Form 63 XLSX template using a per-template column mapping.

    Each teacher occupies two consecutive rows (semester 1, semester 2).
    Cells are merged where appropriate so the visual layout matches the
    target form: row number / ФИО / Должность span both rows, and the
    horizontally-merged header spans (e.g. D:F for ФИО, G:I for Должность)
    are reused for the data rows.
    """
    if not column_mapping:
        raise ValueError("column_mapping пустой — невозможно сгенерировать Форму 63")

    wb = load_workbook(template_path)
    ws = wb.active

    # group input rows by teacher
    grouped: dict[tuple, dict] = {}
    for row in rows:
        key = (row["teacher_name"], row["position"])
        if key not in grouped:
            grouped[key] = {
                "teacher_name": row["teacher_name"],
                "position": row["position"],
                "sem1": None,
                "sem2": None,
            }
        if row["semester"] == 1:
            grouped[key]["sem1"] = row
        elif row["semester"] == 2:
            grouped[key]["sem2"] = row

    teacher_name_span = _category_span_in_header(
        ws, column_mapping, "teacher_name", data_start_row - 4,
    ) or _category_span_in_header(
        ws, column_mapping, "teacher_name", data_start_row - 1,
    )
    position_span = _category_span_in_header(
        ws, column_mapping, "position", data_start_row - 4,
    ) or _category_span_in_header(
        ws, column_mapping, "position", data_start_row - 1,
    )

    teaching_auditory_letter = column_mapping.get("teaching_auditory")
    social_letter = column_mapping.get("social")
    total_letter = column_mapping.get("total")

    def empty_sem(sem_no: int) -> dict:
        return {
            "semester": sem_no,
            "teaching_auditory_hours": 0,
            "teaching_extraauditory_hours": 0,
            "methodical_hours": 0,
            "research_hours": 0,
            "organizational_methodical_hours": 0,
            "educational_hours": 0,
            "qualification_hours": 0,
            "social_hours": 0,
            "planned_total_hours": 0,
            "hourly_auditory_hours": 0,
            "hourly_extraauditory_hours": 0,
        }

    current_row = data_start_row
    no = 1

    for _, teacher in grouped.items():
        sem1 = teacher["sem1"] or empty_sem(1)
        sem2 = teacher["sem2"] or empty_sem(2)

        for offset, sem_data, sem_no in (
            (0, sem1, 1),
            (1, sem2, 2),
        ):
            r = current_row + offset
            _put(ws, column_mapping, "semester", r, sem_no)

            for category, field in DATA_CATEGORY_TO_FIELD.items():
                _put(ws, column_mapping, category, r, sem_data.get(field, 0))

            # Total as a SUM formula across the data columns from
            # teaching_auditory to social (the K..R range in the canonical layout).
            if total_letter and teaching_auditory_letter and social_letter:
                ws[f"{total_letter}{r}"] = (
                    f"=SUM({teaching_auditory_letter}{r}:{social_letter}{r})"
                )

        _put(ws, column_mapping, "row_number", current_row, no)
        _put(ws, column_mapping, "teacher_name", current_row, teacher["teacher_name"])
        _put(ws, column_mapping, "position", current_row, teacher["position"])

        # --- merges for the teacher block (idempotent) ---
        rn_letter = column_mapping.get("row_number")
        if rn_letter:
            _safe_merge(ws, f"{rn_letter}{current_row}:{rn_letter}{current_row + 1}")

        tn_letter = column_mapping.get("teacher_name")
        if tn_letter:
            tn_min, tn_max = (teacher_name_span or (tn_letter, tn_letter))
            _safe_merge(ws, f"{tn_min}{current_row}:{tn_max}{current_row + 1}")

        pos_letter = column_mapping.get("position")
        if pos_letter:
            p_min, p_max = (position_span or (pos_letter, pos_letter))
            _safe_merge(ws, f"{p_min}{current_row}:{p_max}{current_row + 1}")

        current_row += 2
        no += 1

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
